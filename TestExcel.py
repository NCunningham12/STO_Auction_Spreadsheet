from openpyxl import Workbook, load_workbook

wb = load_workbook('PythonTest.xlsx')
ws = wb.active


ws.title = "Initial Python Test"

print("WB Updated")

treeData = [["Make", "Model", "Year", "Mileage", "Purchase Price", "Condition", "BB Value", "Flip Price", "Profit Margin", "Margin %"], ["Toyota", "Model", 2020, 30000, 40000, 3, "BB Value"]]

# Column Titles
# ws['A1'] = 'Make'
# ws['B1'] = 'Model'
# ws['C1'] = 'Year'
# ws['D1'] = 'Mileage'
# ws['E1'] = 'Purchase Price'
# ws['F1'] = 'Condition'
# ws['G1'] = 'BB Value'
# ws['H1'] = 'Flip Price'
# ws['I1'] = 'Profit Margin'
# ws['J1'] = 'Margin %'

for row in treeData:
  ws.append(row)

# Values
ws['A2'] = 'Toyota'
ws['B2'] = 'Tacoma'
ws['C2'] = 2020
ws['D2'] = 30000
ws['E2'] = 40000
ws['F2'] = 3
ws['G2'] = 50000

def Flip():
  original_price = ws['E2'].value
  flip_price = original_price
  bb_value = ws['G2'].value

  if ws['C2'].value >= 2020:
   flip_price += 2000
  
  if ws['D2'].value <= 50000:
    flip_price += 5000

  if ws['F2'].value == 3:
    flip_price += 20000
  elif ws['F2'].value == 2:
    flip_price += 10000

  prof_marg = flip_price - bb_value
  ws['I2'] = prof_marg

  prof_perc = (prof_marg / original_price) * 100

  ws['J2'] = prof_perc

  ws['H2'] = flip_price

Flip()

wb.save('PythonTest.xlsx')