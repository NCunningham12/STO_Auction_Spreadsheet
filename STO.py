from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import html.parser
import requests

# Scrape current Exchange rate (In Development)
# exchange_page = requests.get("https://www.google.com/finance/quote/CAD-USD?hl=en")
# soup = BeautifulSoup(exchange_page.text, "html.parser")
# rate = soup.find("div", attrs={"class":"YM1Kec fxKbKc"})

# print(rate)

# Populate the Workbook
wb = load_workbook('STO_BOB.xlsx')
ws = wb.active

ws.title = "STO LLC Auction Calculations"

truckCategories = [["ID", "Auction Price", "Province", "Canadian Price (CAD)", "US Price (USD)", "Book (USD)", "Total US (USD)", "BOB"]]

# for row in truckCategories:
#   ws.append(row)

# INPUTS: Auction Price, Province, US Price, Book
# OUTPUTS: ID, Canadian Price, Total US, BOB

def BOB(): 
  id = 0
  
  for row in range(2, ws.max_row + 1):
    id += 1
    ws['A' + str(row)].value = id                                     # Creates a unique ID number for each vehicle for reference only

    auction_price = ws['B' + str(row)].value
    province = ws['C' + str(row)].value
    book_usd = ws['D' + str(row)].value
    cad_usd_exchange = ws['K2'].value
    usd_cad_exchange = ws['L2'].value

    book_cad = usd_cad_exchange * book_usd
    ws['E' + str(row)].value = round(book_cad, 2)


    if province.lower() == 'alb' or province.lower() == 'bc':         # Checks if province is WEST
      auction_price += 1000
      auction_price *= 1.05                                           # Adds 1000 and then adds 5% of the new total
      canadian_price = round(auction_price, 2)
      ws['F' + str(row)] = round(canadian_price, 2)

      us_price = round(canadian_price * cad_usd_exchange, 2)           # Converts the CAD to USD based on the input "Exchange rate"
      ws['G' + str(row)] = round(us_price, 2)

      us_price += 2500                                                # Adds 2000 and then adds 1% of the new total
      total_us_price = round(us_price, 2)  
      ws['H' + str(row)] = total_us_price                       
    elif province.lower() == 'ont':                                   # Checks if province is EAST
      auction_price += 800
      auction_price *= 1.13                                           # Adds 1000 and then adds 13% of the new total
      canadian_price = round(auction_price, 2)                        
      ws['F' + str(row)] = round(canadian_price, 2)

      us_price = round(canadian_price * cad_usd_exchange, 2)           # Converts the CAD to USD based on the input "Exchange Rate"
      ws['G' + str(row)] = round(us_price, 2)

      us_price += 1400                                                # Adds 1400 and then adds 1% of the new total
      total_us_price = round(us_price, 2)
      ws['H' + str(row)] = total_us_price

    bob = round(book_usd - total_us_price, 2)                         # Subtracts Total US from Book to calculate BOB
    ws['I' + str(row)] = bob


    # print("Auction Price: " + str(round(auction_price, 2)))
    # print("Canadian Price: " + str(canadian_price))
    # print("US Price: " + str(round(us_price, 2)))
    # print("Total US Price: " + str(total_us_price))
    # print("BOB: " + str(bob))

BOB()

wb.save('STO_BOB.xlsx')
print("✅ WORKBOOK UPDATED")